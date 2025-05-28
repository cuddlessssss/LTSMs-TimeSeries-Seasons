import pandas as pd
import numpy as np
import time
from sklearn.preprocessing import LabelEncoder
import tensorflow as tf
from tensorflow.keras.models import Model
from tensorflow.keras.layers import LSTM, Dense, Input, ReLU
from tensorflow.keras.optimizers import Adam
from tensorflow.keras.callbacks import EarlyStopping
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 🕒 Start timer
total_start = time.time()

# 1. Load & Prepare Future Totals First
future_total_raw = pd.read_excel('future_total_sell_outs_pivoted.xlsx')
future_total_melted = future_total_raw.melt(
    id_vars='Model Name', var_name='year_month', value_name='total_sell_out'
)
future_total_melted['year_month'] = pd.to_datetime(future_total_melted['year_month']).dt.to_period('M')

# 2. Load Past Sales Data
data = pd.read_excel('your_past_data.xlsx', sheet_name='updated')
data.columns = data.columns.str.strip()
data = data[data['Category Description'].str.strip() == 'LCD TV']
data = data[data['Dealer Net Sell Out Qty'] >= 0]
data['Date'] = pd.to_datetime(data['Date'], errors='coerce')
data['year_month'] = data['Date'].dt.to_period('M')
data = data[data['Date'] >= data['Date'].max() - pd.DateOffset(years=4)]

# Label encode
f1_encoder = LabelEncoder()
f2_encoder = LabelEncoder()
data['Model Name_enc'] = f1_encoder.fit_transform(data['Model Name'])
data['Account Name_enc'] = f2_encoder.fit_transform(data['Account Name'])

# Use encoders on future totals
future_total_melted['model_enc'] = f1_encoder.transform(future_total_melted['Model Name'])

# 3. Preprocess Time Series Data
grouped = data.groupby(['year_month', 'Model Name_enc', 'Account Name_enc'])['Dealer Net Sell Out Qty'].sum().reset_index()
pivot = grouped.pivot_table(index='year_month', columns=['Model Name_enc', 'Account Name_enc'], values='Dealer Net Sell Out Qty', fill_value=0)
pivot = pivot.sort_index(axis=1)
column_index = pivot.columns.to_list()

pivot_scaled = pivot.div(pivot.sum(axis=1), axis=0)
target_values = pivot.values

# Add time features (month sin/cos)
months = pivot.index.to_timestamp().month.values
month_sin = np.sin(2 * np.pi * months / 12)
month_cos = np.cos(2 * np.pi * months / 12)
time_features = np.stack([month_sin, month_cos], axis=1)

# 4. Build Input Sequences
input_len = 24
output_len = 12
X_seq, X_time, y_target = [], [], []

for i in range(len(pivot) - input_len - output_len + 1):
    X_seq.append(pivot_scaled.iloc[i:i+input_len].values)
    X_time.append(time_features[i:i+input_len])
    y_target.append(target_values[i+input_len:i+input_len+output_len])

X_seq = np.array(X_seq)
X_time = np.array(X_time)
y_target = np.array(y_target)

X_time_expanded = np.repeat(X_time[:, :, np.newaxis, :], X_seq.shape[2], axis=2)
X_time_expanded = np.transpose(X_time_expanded, (0, 2, 1, 3))
X_seq_expanded = np.transpose(X_seq, (0, 2, 1))[:, :, :, np.newaxis]
X_combined = np.concatenate([X_seq_expanded, X_time_expanded], axis=-1)
X_combined = np.transpose(X_combined, (0, 2, 1, 3)).reshape(X_seq.shape[0], input_len, -1)

# 5. Build & Train Model
training_start = time.time()

input_layer = Input(shape=(input_len, X_combined.shape[2]))
x = LSTM(128, return_sequences=False)(input_layer)
x = Dense(256, activation='softplus')(x)
x = Dense(y_target.shape[1] * y_target.shape[2], activation='softplus')(x)  # Softplus ensures positivity
output_layer = tf.keras.layers.Reshape((y_target.shape[1], y_target.shape[2]))(x)

model = Model(inputs=input_layer, outputs=output_layer)
model.compile(optimizer=Adam(0.001), loss='mse')

model.fit(
    X_combined, y_target,
    epochs=300,
    batch_size=16,
    callbacks=[EarlyStopping(patience=50, restore_best_weights=True)],
    verbose=2
)

print(f"🧠 Training time: {time.time() - training_start:.2f} seconds")

# 6. Forecasting
last_input_seq = pivot_scaled.iloc[-input_len:].values
last_time_feats = time_features[-input_len:]

last_time_exp = np.repeat(last_time_feats[np.newaxis, :, :], last_input_seq.shape[1], axis=0)
last_seq_exp = last_input_seq.T[:, :, np.newaxis]

last_combined = np.concatenate([last_seq_exp, last_time_exp], axis=-1)
last_combined = last_combined.transpose(1, 0, 2).reshape(1, input_len, -1)

forecast_preds = model.predict(last_combined, verbose=0)[0]

# 7. Rescaling & Rounding
def assign_up_flags(group):
    group = group.copy()
    total_target = group['total_model_qty'].iloc[0]
    group_sum = group['rounded_qty'].sum()
    remainder = int(total_target - group_sum)
    group['highlight'] = 'down'
    if remainder > 0:
        group = group.sort_values('residual', ascending=False).copy()
        idxs = group.head(remainder).index
        group.loc[idxs, 'rounded_qty'] += 1
        group.loc[idxs, 'highlight'] = 'up'
    return group

rescaled_forecast = []
for month_idx in range(forecast_preds.shape[0]):
    forecast_month = pivot.index[-1] + (month_idx + 1)
    mask = future_total_melted['year_month'] == forecast_month
    future_totals_for_month = future_total_melted[mask].set_index('model_enc')['total_sell_out'].to_dict()

    for model_enc in f1_encoder.transform(f1_encoder.classes_):
        idxs = [i for i, (m, a) in enumerate(column_index) if m == model_enc]
        preds_for_model = forecast_preds[month_idx, idxs]
        preds_for_model = np.clip(preds_for_model, 0, None)

        sum_preds = preds_for_model.sum()
        total_model_qty = future_totals_for_month.get(model_enc, 0)
        rescale_factor = total_model_qty / sum_preds if sum_preds > 0 else 0

        scaled_preds = preds_for_model * rescale_factor
        residuals = scaled_preds - np.floor(scaled_preds)
        rounded_qty = np.floor(scaled_preds).astype(int)

        df_rounding = pd.DataFrame({
            'index': idxs,
            'residual': residuals,
            'rounded_qty': rounded_qty,
            'total_model_qty': total_model_qty
        })
        df_rounding = assign_up_flags(df_rounding)

        for _, row in df_rounding.iterrows():
            rescaled_forecast.append({
                'year_month': forecast_month,
                'model_enc': model_enc,
                'account_enc': column_index[row['index']][1],
                'rounded_qty': row['rounded_qty'],
                'highlight': row['highlight']
            })

rescaled_df = pd.DataFrame(rescaled_forecast)
rescaled_df['Model Name'] = f1_encoder.inverse_transform(rescaled_df['model_enc'])
rescaled_df['Account Name'] = f2_encoder.inverse_transform(rescaled_df['account_enc'])
rescaled_df['year_month'] = rescaled_df['year_month'].dt.to_timestamp()

# 8. Excel Export
export_pivot = rescaled_df.pivot_table(index=['Model Name', 'Account Name'], columns='year_month', values='rounded_qty', fill_value=0)
highlight_pivot = rescaled_df.pivot_table(index=['Model Name', 'Account Name'], columns='year_month', values='highlight', aggfunc='first', fill_value='down')

wb = Workbook()
ws = wb.active
ws.title = "Forecasted Sell Out"

headers = ['Model Name', 'Account Name'] + [dt.strftime('%Y-%m') for dt in export_pivot.columns]
ws.append(headers)

fill_up = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
fill_down = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')

for (model, account), row in export_pivot.iterrows():
    data_row = [model, account] + row.tolist()
    ws.append(data_row)
    current_row = ws.max_row
    for col_idx, dt in enumerate(export_pivot.columns, start=3):
        highlight = highlight_pivot.loc[(model, account), dt]
        if highlight == 'up':
            ws.cell(row=current_row, column=col_idx).fill = fill_up
        elif highlight == 'down':
            ws.cell(row=current_row, column=col_idx).fill = fill_down

wb.save('forecast_output1.xlsx')
print(f"✅ Finished in {time.time() - total_start:.2f} seconds")
