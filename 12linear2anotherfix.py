import pandas as pd
import numpy as np
import time
from sklearn.preprocessing import LabelEncoder
from sklearn.linear_model import LinearRegression
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 🕒 Start timer
total_start = time.time()

# 1. Load & Prepare Future Totals First
future_total_raw = pd.read_excel('future_total_sell_outs_pivoted.xlsx')
future_total_melted = future_total_raw.melt(
    id_vars='Model Name', var_name='year_month', value_name='total_sell_out'
)
future_total_melted['year_month'] = pd.to_datetime(future_total_melted['year_month']).dt.to_period('M')

# 2. Load Past Sales Data
data = pd.read_excel('your_past_data.xlsx', sheet_name='updated')
data.columns = data.columns.str.strip()
data = data[data['Category Description'].str.strip() == 'LCD TV']
data = data[data['Dealer Net Sell Out Qty'] >= 0]
data['Date'] = pd.to_datetime(data['Date'], errors='coerce')
data['year_month'] = data['Date'].dt.to_period('M')
data = data[data['Date'] >= data['Date'].max() - pd.DateOffset(months=10)]

# Label encode
f1_encoder = LabelEncoder()
f2_encoder = LabelEncoder()
data['Model Name_enc'] = f1_encoder.fit_transform(data['Model Name'])
data['Account Name_enc'] = f2_encoder.fit_transform(data['Account Name'])

# Use encoders on future totals
future_total_melted['model_enc'] = f1_encoder.transform(future_total_melted['Model Name'])

# 3. Preprocess Time Series Data
grouped = data.groupby(['year_month', 'Model Name_enc', 'Account Name_enc'])['Dealer Net Sell Out Qty'].sum().reset_index()
pivot = grouped.pivot_table(index='year_month', columns=['Model Name_enc', 'Account Name_enc'], values='Dealer Net Sell Out Qty', fill_value=0)
pivot = pivot.sort_index(axis=1)
column_index = pivot.columns.to_list()

# 4. Calculate trend + seasonality using linear regression with per-account month dummies
trend_preds = {}

for col_idx, (model_enc, account_enc) in enumerate(pivot.columns):
    y = pivot.iloc[:, col_idx].values

    # Remove extreme values (outside 1st and 99th percentiles)
    p1, p99 = np.percentile(y, [1, 99])
    mask = (y >= p1) & (y <= p99)

    months_numeric = np.arange(len(pivot.index)).reshape(-1, 1)
    account_months = pivot.index.to_timestamp().month
    account_month_dummies = pd.get_dummies(account_months)
    X_trend_season = np.hstack([months_numeric, account_month_dummies])

    reg = LinearRegression()
    reg.fit(X_trend_season[mask], y[mask])

    future_months = np.arange(len(pivot.index), len(pivot.index) + 12).reshape(-1, 1)
    last_month = pivot.index[-1].to_timestamp()
    future_periods = pd.date_range(start=last_month + pd.offsets.MonthBegin(1), periods=12, freq='ME')
    future_month_nums = future_periods.month
    future_month_dummies = pd.get_dummies(future_month_nums)
    future_month_dummies = future_month_dummies.reindex(columns=account_month_dummies.columns, fill_value=0)
    X_future = np.hstack([future_months, future_month_dummies])

    future_vals = reg.predict(X_future)
    trend_preds[(model_enc, account_enc)] = future_vals

# 5. Assemble trend_preds into forecast_preds array
forecast_preds = np.zeros((12, len(pivot.columns)))
for idx, key in enumerate(pivot.columns):
    forecast_preds[:, idx] = trend_preds[key]

# 6. Rescaling & Rounding
def assign_up_flags(group):
    group = group.copy()
    total_target = group['total_model_qty'].iloc[0]
    group_sum = group['rounded_qty'].sum()
    remainder = int(total_target - group_sum)
    group['highlight'] = 'down'
    if remainder > 0:
        group = group.sort_values('residual', ascending=False).copy()
        idxs = group.head(remainder).index
        group.loc[idxs, 'rounded_qty'] += 1
        group.loc[idxs, 'highlight'] = 'up'
    return group

rescaled_forecast = []
for month_idx in range(forecast_preds.shape[0]):
    forecast_month = pivot.index[-1] + (month_idx + 1)
    mask = future_total_melted['year_month'] == forecast_month
    future_totals_for_month = future_total_melted[mask].set_index('model_enc')['total_sell_out'].to_dict()

    for model_enc in f1_encoder.transform(f1_encoder.classes_):
        idxs = [i for i, (m, a) in enumerate(column_index) if m == model_enc]
        preds_for_model = forecast_preds[month_idx, idxs]

        preds_for_model = np.maximum(preds_for_model, 0)

        sum_preds = preds_for_model.sum()
        total_model_qty = future_totals_for_month.get(model_enc, 0)
        rescale_factor = total_model_qty / sum_preds if sum_preds > 0 else 0

        scaled_preds = preds_for_model * rescale_factor
        residuals = scaled_preds - np.floor(scaled_preds)
        rounded_qty = np.floor(scaled_preds).astype(int)

        df_rounding = pd.DataFrame({
            'index': idxs,
            'residual': residuals,
            'rounded_qty': rounded_qty,
            'total_model_qty': total_model_qty
        })
        df_rounding = assign_up_flags(df_rounding)

        for _, row in df_rounding.iterrows():
            rescaled_forecast.append({
                'year_month': forecast_month,
                'model_enc': model_enc,
                'account_enc': column_index[row['index']][1],
                'rounded_qty': row['rounded_qty'],
                'highlight': row['highlight']
            })

rescaled_df = pd.DataFrame(rescaled_forecast)
rescaled_df['Model Name'] = f1_encoder.inverse_transform(rescaled_df['model_enc'])
rescaled_df['Account Name'] = f2_encoder.inverse_transform(rescaled_df['account_enc'])
rescaled_df['year_month'] = rescaled_df['year_month'].dt.to_timestamp()

# 7. Excel Export
export_pivot = rescaled_df.pivot_table(index=['Model Name', 'Account Name'], columns='year_month', values='rounded_qty', fill_value=0)
highlight_pivot = rescaled_df.pivot_table(index=['Model Name', 'Account Name'], columns='year_month', values='highlight', aggfunc='first', fill_value='down')

wb = Workbook()
ws = wb.active
ws.title = "Forecasted Sell Out"

headers = ['Model Name', 'Account Name'] + [dt.strftime('%Y-%m') for dt in export_pivot.columns]
ws.append(headers)

fill_up = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
fill_down = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')

for (model, account), row in export_pivot.iterrows():
    data_row = [model, account] + row.tolist()
    ws.append(data_row)
    current_row = ws.max_row
    for col_idx, dt in enumerate(export_pivot.columns, start=3):
        highlight = highlight_pivot.loc[(model, account), dt]
        if highlight == 'up':
            ws.cell(row=current_row, column=col_idx).fill = fill_up
        elif highlight == 'down':
            ws.cell(row=current_row, column=col_idx).fill = fill_down

wb.save('forecast_output_trend_seasonality_testing_actuals2.xlsx')
print(f"✅ Finished in {time.time() - total_start:.2f} seconds")
